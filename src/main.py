import json
import os
import csv
from openpyxl import load_workbook
from masks import mask_card_number, mask_account
from utils import search_transactions, categorize_transactions

def load_json(filename):
    """Загрузка данных из JSON-файла."""
    full_path = os.path.join(os.path.dirname(__file__), "..", "data", filename)
    with open(full_path, "r", encoding="utf-8") as f:
        return json.load(f)

def load_csv(filename):
    """Загрузка данных из CSV-файла."""
    full_path = os.path.join(os.path.dirname(__file__), "..", "data", filename)
    transactions = []
    with open(full_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            transactions.append(row)
    return transactions

def load_xlsx(filename):
    """Загрузка данных из XLSX-файла."""
    full_path = os.path.join(os.path.dirname(__file__), "..", "data", filename)
    workbook = load_workbook(full_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    transactions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        transaction = {headers[i]: row[i] for i in range(len(headers))}
        transactions.append(transaction)
    return transactions

def load_data(filename):
    """Универсальная функция загрузки данных."""
    _, ext = os.path.splitext(filename)
    if ext.lower() == ".json":
        return load_json(filename)
    elif ext.lower() == ".csv":
        return load_csv(filename)
    elif ext.lower() == ".xlsx":
        return load_xlsx(filename)
    else:
        raise ValueError("Неподдерживаемый формат файла")

def filter_by_status(transactions, status):
    return [t for t in transactions if t.get("state", "").upper() == status.upper()]

def sort_by_date(transactions, reverse=False):
    from datetime import datetime
    def parse_date(t):
        return datetime.fromisoformat(t.get("date", ""))
    return sorted(transactions, key=parse_date, reverse=reverse)

def filter_by_currency(transactions, currency):
    return [
        t for t in transactions
        if t.get("operationAmount", {}).get("currency", {}).get("code") == currency
    ]

def main():
    print("Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    print("Выберите необходимый пункт меню:")
    print("1. Получить информацию о транзакциях из JSON-файла")
    print("2. Получить информацию о транзакциях из CSV-файла")
    print("3. Получить информацию о транзакциях из XLSX-файла")

    choice = input("\nПользователь: ").strip()
    if choice not in ["1", "2", "3"]:
        print("Неверный выбор. Попробуйте снова.\n")
        return

    try:
        if choice == "1":
            filename = input("Введите имя JSON-файла: ").strip()
        elif choice == "2":
            filename = input("Введите имя CSV-файла: ").strip()
        elif choice == "3":
            filename = input("Введите имя XLSX-файла: ").strip()

        transactions = load_data(filename)
        print(f"Для обработки выбран файл: {filename}")

        # Фильтрация по статусу
        while True:
            print("\nВведите статус, по которому необходимо выполнить фильтрацию.")
            print("Доступные для фильтровки статусы: EXECUTED, CANCELED, PENDING")
            status = input("Пользователь: ").strip().upper()
            if status not in ["EXECUTED", "CANCELED", "PENDING"]:
                print(f'Статус операции "{status}" недоступен.')
                continue
            filtered = filter_by_status(transactions, status)
            print(f'Операции отфильтрованы по статусу "{status}"')
            break

        # Сортировка по дате
        sort_choice = input("\nОтсортировать операции по дате? Да/Нет\nПользователь: ").strip().lower()
        if sort_choice == "да":
            order = input("Отсортировать по возрастанию или по убыванию?\nПользователь: ").strip().lower()
            filtered = sort_by_date(filtered, reverse="убыв" in order)
            print("Операции отсортированы по дате.")

        # Фильтр по валюте
        rub_filter = input("\nВыводить только рублёвые транзакции? Да/Нет\nПользователь: ").strip().lower()
        if rub_filter == "да":
            filtered = filter_by_currency(filtered, "RUB")
            print("Фильтрация по рублям завершена.")

        # Поиск по описанию
        search_choice = input("\nОтфильтровать список транзакций по слову в описании? Да/Нет\nПользователь: ").strip().lower()
        if search_choice == "да":
            keyword = input("Введите ключевое слово:\nПользователь: ").strip()
            filtered = search_transactions(filtered, keyword)
            print("Фильтрация по ключевому слову завершена.")

        if not filtered:
            print("\nНе найдено ни одной транзакции, подходящей под ваши условия фильтрации.")
            return

        print("\nРаспечатываю итоговый список транзакций...")
        print(f"\nВсего банковских операций в выборке: {len(filtered)}\n")

        for t in filtered:
            description = t["description"]
            amount = t["operationAmount"]["amount"]
            currency = t["operationAmount"]["currency"]["code"]

            from_acc = t.get("from", "Не указан")
            to_acc = t.get("to", "Не указан")

            if "Счет" in from_acc:
                from_acc = mask_account(from_acc)
            else:
                from_acc = mask_card_number(from_acc)

            to_acc = mask_account(to_acc)

            print(f"{description}")
            print(f"{from_acc} -> {to_acc}")
            print(f"Сумма: {amount} {currency}\n")

    except FileNotFoundError:
        print("Файл данных не найден. Проверьте путь.")
    except Exception as e:
        print(f"Ошибка при выполнении программы: {e}")

if __name__ == "__main__":
    main()